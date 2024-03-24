from typing import Dict, Optional, List, Tuple, Any
import openpyxl
import os

from core.bot import Bot
from core.data import DataChunkBase, custom_data_chunk
from core.command.const import *
from core.command import UserCommandBase, custom_user_command
from core.command import BotCommandBase, BotSendMsgCommand
from core.communication import MessageMetaData, PrivateMessagePort, GroupMessagePort
from core.config import DATA_PATH
from module.common import DC_GROUPCONFIG
from core.localization import LOC_PERMISSION_DENIED_NOTICE, LOC_FUNC_DISABLE

LOC_MODE_SWITCH = "mode_switch"
LOC_MODE_INVALID = "mode_invalid"
LOC_MODE_NOT_EXIST = "mode_not_exist"
LOC_MODE_LIST = "mode_list"
LOC_MODE_LIKELY = "mode_likely"

CFG_MODE_ENABLE = "mode_enable"
CFG_MODE_DEFAULT = "mode_default"

MODE_FILE_PATH = "Config/mode_setting.xlsx"

DEFAULT_FIELD = ['mode','default_dice', 'query_database']
DEFAULT_TABLE = [
    ["DND5E", "20", "DND5E"],
    ["PF1E", "20", "PF1E"],
    ["COC7", "100", "COC7"],
    ["NECHRONICA", "10", "NECHRONICA"],
]

@custom_user_command(readable_name="模式指令", priority=-2,
                     flag=DPP_COMMAND_FLAG_MANAGE, group_only=True
                     )
class ModeCommand(UserCommandBase):
    """
    .mode 模式设置指令（批量群设置修改/模板调用指令）
    """

    def __init__(self, bot: Bot):
        super().__init__(bot)
        bot.loc_helper.register_loc_text(LOC_MODE_SWITCH, "已切换至{new_mode}模式（默认{dice}面骰点，查询数据库使用{database}.db（如果有））。", "。mode切换群模式指令，切换模式等于一次性修改多个群配置。\nnew_mode：切换后的模式，dice：默认骰面，database：查询使用数据库")
        bot.loc_helper.register_loc_text(LOC_MODE_INVALID, "该模式配置有误，无法切换，请询问骰主。", "。mode切换群模式，但模式文件有问题的情况下返回")
        bot.loc_helper.register_loc_text(LOC_MODE_NOT_EXIST, "该模式不存在！", "。mode切换群模式，但模式文件中不存在此模式时返回")
        bot.loc_helper.register_loc_text(LOC_MODE_LIST, "以下是可用的模式列表：{modes}", "。mode模式指令查看可用模式列表\nmodes：可用模式列表")
        bot.loc_helper.register_loc_text(LOC_MODE_LIKELY, "找到多个选项，你要找的是不是：{modes}", "。mode模式指令，模糊匹配出现多个结果\nmodes：模糊匹配结果列表")

        bot.cfg_helper.register_config(CFG_MODE_ENABLE, "1", "模式指令开关")
        bot.cfg_helper.register_config(CFG_MODE_DEFAULT, "DND5E", "群内默认模式")

        self.mode_dict: Dict[str,List[str]] = {}
        self.mode_field: List[str] = DEFAULT_FIELD

    def delay_init(self) -> List[str]:
        bot_id: str = self.bot.account
        init_info: List[str] = []
        edited: bool = False
        # 从本地文件中读取可用模式一览
        data_path = os.path.join(DATA_PATH, MODE_FILE_PATH)
        if os.path.exists(data_path):
            wb = openpyxl.load_workbook(data_path)
            id_list = wb.get_sheet_names()
            if bot_id in id_list:
                ws = wb.get_sheet_by_name(bot_id)
                for row in ws:
                    if str(row[0].value) == "mode":
                        self.mode_field = [str(cell.value) for cell in row]
                    else:
                        self.mode_dict[str(row[0].value)] = [str(cell.value) for cell in row[1:]]
            else:
                ws = wb.create_sheet(bot_id)
                ws.append(DEFAULT_FIELD)
                for str_row in DEFAULT_TABLE:
                    ws.append(str_row)
                    self.mode_dict[str_row[0]] = [str_cell for str_cell in str_row[1:]]
                edited = True
            init_info.append("已载入模式文件。")
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = bot_id
            ws.append(DEFAULT_FIELD)
            for str_row in DEFAULT_TABLE:
                ws.append(str_row)
                self.mode_dict[str_row[0]] = [str_cell for str_cell in str_row[1:]]
            edited = True
            init_info.append("已创建模式文件。")
        if edited:
            wb.save(data_path)
        return init_info

    def can_process_msg(self, msg_str: str, meta: MessageMetaData) -> Tuple[bool, bool, Any]:
        should_proc: bool = True
        should_pass: bool = False
        hint: str = ""
        # 判断是否初始化，没有初始化则进行一次初始化
        if self.bot.data_manager.get_data(DC_GROUPCONFIG,[meta.group_id,"mode"],default_val="") == "":
            default_mode = str(self.bot.cfg_helper.get_config(CFG_MODE_DEFAULT)[0])
            if default_mode != "":
                self.switch_mode(meta.group_id,default_mode)
            else:
                self.bot.data_manager.set_data(DC_GROUPCONFIG,[meta.group_id,"mode"],"NULL")
        # 判断指令
        if not meta.group_id:
            should_proc = False
        elif msg_str.startswith(".模式"):
            hint = msg_str[3:].strip()
        elif msg_str.startswith(".mode"):
            hint = msg_str[5:].strip()
        else:
            should_proc = False
        return should_proc, should_pass, hint

    def process_msg(self, msg_str: str, meta: MessageMetaData, hint: Any) -> List[BotCommandBase]:
        port = GroupMessagePort(meta.group_id) if meta.group_id else PrivateMessagePort(meta.user_id)
        # 判断功能开关
        try:
            assert (int(self.bot.cfg_helper.get_config(CFG_MODE_ENABLE)[0]) != 0)
        except AssertionError:
            feedback = self.bot.loc_helper.format_loc_text(LOC_FUNC_DISABLE, func=self.readable_name)
            return [BotSendMsgCommand(self.bot.account, feedback, [port])]
        # 判断权限
        if meta.permission < 1: # 需要至少1级权限（群管理/骰管理）才能执行
            feedback = self.bot.loc_helper.format_loc_text(LOC_PERMISSION_DENIED_NOTICE)
            return [BotSendMsgCommand(self.bot.account, feedback, [port])]
        # 解析语句
        arg_var = hint.strip().upper()

        if arg_var == "DEFAULT" or arg_var == "CLEAR":
            feedback = self.switch_mode(meta.group_id,self.bot.cfg_helper.get_config(CFG_MODE_DEFAULT)[0])
        elif arg_var != "":
            feedback = self.switch_mode(meta.group_id,arg_var)
        else:
            feedback = self.bot.loc_helper.format_loc_text(LOC_MODE_LIST, modes = "、".join(self.mode_dict.keys()))

        return [BotSendMsgCommand(self.bot.account, feedback, [port])]
    
    def switch_mode(self, group_id: str, mode: str) -> str:
        # 居然不能引用，只能在这再搭一个了
        def update_group_config(group_id: str, setting: List[str], var:List[str]):
            self.bot.data_manager.delete_data(DC_GROUPCONFIG, [group_id])
            for index in range(len(setting)):
                true_var: Any
                if var[index].isdigit():
                    true_var = int(var[index])
                elif var[index].upper() == "TRUE":
                    true_var = True
                elif var[index].upper() == "FALSE":
                    true_var = False
                else:
                    true_var = var[index]
                self.bot.data_manager.set_data(DC_GROUPCONFIG, [group_id,setting[index]],true_var)
        
        matched = False
        feedback = ""
        # 尝试精准匹配
        for key in self.mode_dict.keys():
            key = key.upper()
            if key.upper() == mode: #精准匹配
                update_group_config(group_id,self.mode_field,[key]+self.mode_dict[key])
                feedback = self.bot.loc_helper.format_loc_text(LOC_MODE_SWITCH, new_mode = key, dice=self.mode_dict[key][0], database=self.mode_dict[key][1])
                matched = True
        # 尝试精准匹配
        if not matched:
            result: List[str] = []
            for key in self.mode_dict.keys():
                key = key.upper()
                if mode in key:
                    result.append(key)
            if len(result) > 1:
                feedback = self.bot.loc_helper.format_loc_text(LOC_MODE_LIKELY,modes="、".join(result))
            elif len(result) == 1:
                key = result[0]
                update_group_config(group_id,self.mode_field,[key]+self.mode_dict[key])
                feedback = self.bot.loc_helper.format_loc_text(LOC_MODE_SWITCH, new_mode = key, dice=self.mode_dict[key][0], database=self.mode_dict[key][1])
                matched = True
            else:
                feedback = self.bot.loc_helper.format_loc_text(LOC_MODE_NOT_EXIST)+self.bot.loc_helper.format_loc_text(LOC_MODE_LIST, modes = "、".join(self.mode_dict.keys()))
        
        return feedback

    def get_help(self, keyword: str, meta: MessageMetaData) -> str:
        if keyword == "config" or keyword == "mode":  # help后的接着的内容
            feedback: str = ".mode dnd/coc/ygo" \
                            "套用模式设置" 
            return feedback
        return ""

    def get_description(self) -> str:
        return ".mode 模式系统"  # help指令中返回的内容
